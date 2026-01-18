import os
import sys
import time
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, LEFT, RIGHT, TOP, BOTTOM, BOTH, X, Y, YES, NO, VERTICAL, HORIZONTAL, W, E, N, S, CENTER
import tempfile
import shutil
import json
import subprocess
import winsound
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.widgets.scrolled import ScrolledText
import win32com.client
import pythoncom
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.gridspec as gridspec
import datetime
from copy import copy

# Configuration
def get_base_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    try:
        return os.path.dirname(os.path.abspath(__file__))
    except Exception:
        return os.getcwd()

BASE_DIR = get_base_dir()
DEFAULT_SOURCE_DIR = os.path.join(BASE_DIR, "3-设备缺陷问题库及设备缺陷处理记录")
TARGET_EXCEL_PATH = os.path.join(BASE_DIR, "设备缺陷问题库（日常巡视、故障处理问题库，广供记-002汇总表，202601起）.xlsx")

# Enable High DPI support
try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)
except:
    pass

def _app_state_path():
    return os.path.join(get_base_dir(), ".app_state.json")

def _load_app_state():
    path = _app_state_path()
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    return {}

def _save_app_state(data):
    path = _app_state_path()
    tmp = f"{path}.tmp"
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
        os.replace(tmp, path)
        return True
    except Exception:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass
        return False

class DefectProcessor:
    def __init__(self, log_callback=print, progress_callback=None):
        self.log = log_callback
        self.progress = progress_callback
        self.stop_requested = False
        self.paused = False

    def _is_doc_path_string(self, value):
        if not isinstance(value, str):
            return False
        s = value.strip()
        if not s:
            return False
        if not s.lower().endswith((".doc", ".docx")):
            return False
        return (":\\" in s) or ("\\" in s) or ("/" in s)

    def _safe_temp_name(self, name):
        s = str(name or "")
        for ch in ['\\', '/', ':', '*', '?', '"', '<', '>', '|', '：']:
            s = s.replace(ch, "_")
        s = s.strip()
        return s or "word.doc"

    def _open_word_doc(self, word, file_path):
        open_kwargs = dict(
            ReadOnly=True,
            AddToRecentFiles=False,
            ConfirmConversions=False,
            Visible=False,
            OpenAndRepair=True,
        )
        return word.Documents.Open(file_path, **open_kwargs)

    def _load_processed_paths_from_excel(self, target_excel):
        paths = set()
        try:
            wb = openpyxl.load_workbook(target_excel, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=4, min_col=14, max_col=14, values_only=True):
                v = row[0] if row else None
                if not isinstance(v, str):
                    continue
                s = v.strip()
                if not s:
                    continue
                paths.add(os.path.normcase(os.path.normpath(s)))
            try:
                wb.close()
            except Exception:
                pass
        except Exception:
            return set()
        return paths

    def _row_has_content(self, row_data):
        if not row_data or len(row_data) <= 1:
            return False
        for cell in row_data[1:]:
            if self._is_doc_path_string(cell):
                continue
            if str(cell).strip():
                return True
        return False

    def _coerce_int(self, value, default=0):
        try:
            if value is None:
                return default
            s = str(value).strip()
            if s == "":
                return default
            return int(float(s))
        except Exception:
            return default

    def _find_last_valid_row(self, ws, min_row=3, serial_col=1, max_cols=13):
        for row in range(ws.max_row, min_row - 1, -1):
            serial_val = ws.cell(row=row, column=serial_col).value
            if serial_val is None or str(serial_val).strip() == "":
                continue
            has_any = False
            for c in range(2, max_cols + 1):
                v = ws.cell(row=row, column=c).value
                if self._is_doc_path_string(v):
                    continue
                if v is not None and str(v).strip() != "":
                    has_any = True
                    break
            if has_any:
                return row
        return 0

    def _row_has_any_defect_cells(self, ws, row, start_col=2, end_col=13):
        for c in range(start_col, end_col + 1):
            v = ws.cell(row=row, column=c).value
            if v is not None and str(v).strip() != "":
                return True
        return False

    def _normalize_excel_rows(self, target_excel):
        try:
            wb = openpyxl.load_workbook(target_excel)
            ws = wb.active

            deleted = 0
            for row in range(ws.max_row, 3, -1):
                path_val = ws.cell(row=row, column=14).value
                serial_val = ws.cell(row=row, column=1).value
                has_defect = self._row_has_any_defect_cells(ws, row, start_col=2, end_col=13)
                has_path = isinstance(path_val, str) and path_val.strip() != ""
                has_serial = serial_val is not None and str(serial_val).strip() != ""

                if not has_defect and (has_path or has_serial):
                    ws.delete_rows(row, 1)
                    deleted += 1

            serial = 0
            for row in range(4, ws.max_row + 1):
                if self._row_has_any_defect_cells(ws, row, start_col=2, end_col=13):
                    serial += 1
                    ws.cell(row=row, column=1).value = serial
                else:
                    ws.cell(row=row, column=1).value = None

            try:
                ws.column_dimensions[get_column_letter(14)].hidden = True
            except Exception:
                pass

            wb.save(target_excel)
            return deleted
        except PermissionError:
            raise
        except Exception as e:
            self.log(f"规范化Excel数据时出错: {e}")
            return 0

    def _estimate_row_height(self, row_data, base_height=45, max_height=150):
        data = list(row_data or [])
        if data:
            last = data[-1]
            if isinstance(last, str):
                s = last.strip()
                if s and (":\\" in s or "\\" in s or "/" in s) and s.lower().endswith((".doc", ".docx")):
                    data = data[:-1]
        max_len = 0
        for cell_text in data:
            if cell_text is None:
                continue
            max_len = max(max_len, len(str(cell_text)))
        est_lines = (max_len / 25) + 1
        height = max(base_height, est_lines * 15)
        return min(height, max_height)

    def _apply_template_style(self, dst_cell, src_cell):
        try:
            dst_cell._style = copy(src_cell._style)
        except Exception:
            pass
        try:
            dst_cell.font = copy(src_cell.font)
        except Exception:
            pass
        try:
            dst_cell.border = copy(src_cell.border)
        except Exception:
            pass
        try:
            dst_cell.fill = copy(src_cell.fill)
        except Exception:
            pass
        try:
            dst_cell.number_format = src_cell.number_format
        except Exception:
            pass
        try:
            dst_cell.protection = copy(src_cell.protection)
        except Exception:
            pass
        try:
            base_alignment = copy(src_cell.alignment)
            try:
                dst_cell.alignment = base_alignment.copy(wrapText=True)
            except Exception:
                dst_cell.alignment = Alignment(horizontal=base_alignment.horizontal, vertical=base_alignment.vertical, wrap_text=True)
        except Exception:
            dst_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _write_rows_to_excel(self, target_excel, extracted_rows, overwrite=False):
        wb = openpyxl.load_workbook(target_excel)
        ws = wb.active

        if overwrite:
            template_row = 3 if ws.max_row >= 3 else 1
            if ws.max_row >= 4:
                try:
                    ws.delete_rows(4, ws.max_row - 3)
                except Exception:
                    pass
            last_serial = 0
            current_row = 4
        else:
            last_valid_row = self._find_last_valid_row(ws, min_row=3, serial_col=1, max_cols=14)
            template_row = last_valid_row if last_valid_row >= 3 else 3
            last_serial = self._coerce_int(ws.cell(row=last_valid_row, column=1).value) if last_valid_row >= 3 else 0
            current_row = (last_valid_row if last_valid_row >= 3 else 2) + 1

        template_cells = [ws.cell(row=template_row, column=c) for c in range(1, 14)]
        template_height = ws.row_dimensions[template_row].height
        if template_height is None:
            template_height = 45
        serial = last_serial

        try:
            ws.column_dimensions[get_column_letter(14)].hidden = True
        except Exception:
            pass

        wrote = 0
        for row_data in extracted_rows:
            if not self._row_has_content(row_data):
                continue

            serial += 1
            if len(row_data) < 14:
                row_data = list(row_data) + [""] * (14 - len(row_data))
            row_data = row_data[:14]
            row_data[0] = str(serial)

            height = self._estimate_row_height(row_data, base_height=template_height)
            ws.row_dimensions[current_row].height = height

            for col_idx, value in enumerate(row_data, start=1):
                dst_cell = ws.cell(row=current_row, column=col_idx, value=value)
                if col_idx <= 13:
                    self._apply_template_style(dst_cell, template_cells[col_idx - 1])

            wrote += 1
            current_row += 1

        wb.save(target_excel)
        return wrote

    def _remove_rows_by_paths(self, target_excel, paths_to_remove):
        if not paths_to_remove:
            return 0
        
        try:
            wb = openpyxl.load_workbook(target_excel)
            ws = wb.active
            
            rows_to_delete = []
            
            # Scan all rows to find matches
            # Data starts at row 4
            for row in range(ws.max_row, 3, -1):
                cell_val = ws.cell(row=row, column=14).value
                if cell_val:
                    s_val = str(cell_val).strip()
                    if s_val:
                        norm_val = os.path.normcase(os.path.normpath(s_val))
                        if norm_val in paths_to_remove:
                            rows_to_delete.append(row)
            
            if not rows_to_delete:
                return 0
            
            self.log(f"发现 {len(rows_to_delete)} 条记录对应已删除的文件，正在清理...")
            
            for r in rows_to_delete:
                ws.delete_rows(r, 1)
                
            # Re-serialize
            serial = 0
            for row in range(4, ws.max_row + 1):
                if self._row_has_any_defect_cells(ws, row, start_col=2, end_col=13):
                    serial += 1
                    ws.cell(row=row, column=1).value = serial
                else:
                    ws.cell(row=row, column=1).value = None

            wb.save(target_excel)
            return len(rows_to_delete)
            
        except Exception as e:
            self.log(f"清理删除文件数据时出错: {e}")
            return 0

    def update_single_file(self, file_path, target_excel):
        try:
            pythoncom.CoInitialize()
        except:
            pass
            
        self.log(f"正在更新单个文件: {file_path}")
        if not os.path.exists(file_path):
            self.log(f"文件不存在: {file_path}")
            return False

        extracted_rows = []
        word = None
        doc = None
        try:
            try:
                word = win32com.client.Dispatch("Word.Application")
            except Exception:
                try:
                    word = win32com.client.Dispatch("Kwps.Application")
                except Exception:
                    word = win32com.client.Dispatch("Wps.Application")
            
            try:
                doc = self._open_word_doc(word, file_path)
            except Exception:
                time.sleep(1)
                doc = self._open_word_doc(word, file_path)

            if doc.Tables.Count > 0:
                table = doc.Tables(1)
                row_count = table.Rows.Count
                if row_count > 1:
                    for r in range(2, row_count + 1):
                        row_data = []
                        try:
                            for c in range(1, 14):
                                try:
                                    cell_text = table.Cell(r, c).Range.Text
                                    cell_text = cell_text.replace('\r', '').replace('\x07', '').strip()
                                    row_data.append(cell_text)
                                except Exception:
                                    row_data.append("")
                            
                            has_content = False
                            if len(row_data) > 1:
                                for cell in row_data[1:]:
                                    if cell.strip():
                                        has_content = True
                                        break
                            if has_content:
                                row_data.append(file_path)
                                extracted_rows.append(row_data)
                        except Exception:
                            pass
        except Exception as e:
            self.log(f"读取Word文件失败: {e}")
            return False
        finally:
            if doc:
                try:
                    doc.Close(False)
                except Exception:
                    pass

        try:
            self._remove_rows_by_paths(target_excel, {os.path.normcase(os.path.normpath(file_path))})
            if extracted_rows:
                self._write_rows_to_excel(target_excel, extracted_rows, overwrite=False)
            self._normalize_excel_rows(target_excel)
            self.log("单文件更新完成。")
            return True
        except Exception as e:
            self.log(f"写入Excel失败: {e}")
            return False

    def sync_word_from_excel(self, target_excel):
        self.log("正在读取Excel数据以同步到Word...")

        def clean_word_text(s):
            try:
                s = "" if s is None else str(s)
            except Exception:
                return ""
            return s.replace("\r", "").replace("\x07", "").strip()

        def fmt_excel_value(v):
            if v is None:
                return ""
            try:
                if isinstance(v, datetime.datetime):
                    if v.hour == 0 and v.minute == 0 and v.second == 0:
                        return v.strftime("%Y-%m-%d")
                    return v.strftime("%Y-%m-%d %H:%M:%S")
                if isinstance(v, datetime.date):
                    return v.strftime("%Y-%m-%d")
            except Exception:
                pass
            try:
                s = str(v)
            except Exception:
                return ""
            s = s.strip()
            return "" if s == "None" else s

        def norm_key_part(v):
            s = clean_word_text(v)
            s = " ".join(s.split())
            if s:
                try:
                    ts = pd.to_datetime(s, errors="coerce")
                    if pd.notna(ts) and 2000 <= int(ts.year) <= 2100:
                        if int(ts.hour) == 0 and int(ts.minute) == 0 and int(ts.second) == 0:
                            return ts.strftime("%Y-%m-%d")
                        return ts.strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    pass
            return s

        def pick_headers(ws):
            for r in (3, 2, 1):
                vals = []
                non_empty = 0
                for c in range(1, 14):
                    v = ws.cell(row=r, column=c).value
                    s = "" if v is None else str(v).strip()
                    if s:
                        non_empty += 1
                    vals.append(s)
                if non_empty >= 3:
                    return vals
            return [""] * 13

        def build_key(values, key_cols):
            parts = []
            for idx in key_cols:
                try:
                    parts.append(norm_key_part(values[idx]))
                except Exception:
                    parts.append("")
            parts = [p for p in parts if p]
            return "||".join(parts)

        def open_word_doc_editable(word_app, file_path):
            open_kwargs = dict(
                ReadOnly=False,
                AddToRecentFiles=False,
                ConfirmConversions=False,
                Visible=False,
                OpenAndRepair=True,
            )
            return word_app.Documents.Open(file_path, **open_kwargs)

        try:
            wb = openpyxl.load_workbook(target_excel, data_only=True)
            ws = wb.active
        except Exception as e:
            self.log(f"无法读取Excel文件: {e}")
            return False

        headers = pick_headers(ws)

        update_cols = []
        for i, h in enumerate(headers):
            if not h:
                continue
            if i == 0:
                continue
            if any(k in h for k in ["销号", "处理", "原因", "措施", "整改", "备注", "状态", "完成"]):
                update_cols.append(i)

        update_cols = sorted(set(update_cols))
        if not update_cols:
            self.log("未识别到可同步的字段（如“销号时间/处理情况/备注/状态”等），为避免误写，已取消反向同步。")
            return False

        key_cols = []
        for i, h in enumerate(headers):
            if not h:
                continue
            if any(k in h for k in ["销号", "处理", "原因", "措施", "整改", "备注", "状态", "完成"]):
                continue
            if any(k in h for k in ["描述", "地点", "位置", "类型", "发现", "发生", "时间", "日期", "编号"]):
                key_cols.append(i)

        key_cols = sorted(set(key_cols))
        if not key_cols:
            key_cols = [i for i in range(1, 13) if i not in update_cols][:4]

        base_cols = [i for i in range(1, 13) if i not in update_cols]
        if not base_cols:
            base_cols = key_cols[:]

        file_rows = {}
        total_rows = 0
        for excel_row_idx, row in enumerate(ws.iter_rows(min_row=4, max_col=14, values_only=True), start=4):
            if not row or len(row) < 14:
                continue

            path_val = row[13]
            if not isinstance(path_val, str) or not path_val.strip():
                continue
            file_path = os.path.normpath(path_val.strip())

            values = list(row[:13])
            if len(values) < 13:
                values.extend([None] * (13 - len(values)))
            values = values[:13]

            has_any = False
            for v in values[1:]:
                if v is None:
                    continue
                if str(v).strip() != "":
                    has_any = True
                    break
            if not has_any:
                continue

            primary_key = build_key(values, key_cols)
            sig = build_key(values, base_cols)
            if not primary_key or not sig:
                continue

            file_rows.setdefault(file_path, []).append((excel_row_idx, values, primary_key, sig))
            total_rows += 1

        if not file_rows:
            self.log("Excel中没有可用于反向同步的数据记录。")
            return True

        word_app = None
        try:
            pythoncom.CoInitialize()
        except Exception:
            pass

        try:
            try:
                word_app = win32com.client.DispatchEx("Word.Application")
            except Exception as e:
                try:
                    word_app = win32com.client.DispatchEx("Kwps.Application")
                except Exception:
                    try:
                        word_app = win32com.client.DispatchEx("Wps.Application")
                    except Exception:
                        raise e

            try:
                word_app.Visible = False
            except Exception:
                pass
            try:
                word_app.DisplayAlerts = 0
            except Exception:
                pass
            try:
                word_app.AutomationSecurity = 3
            except Exception:
                pass

            total_files = len(file_rows)
            updated_files = 0
            updated_cells = 0
            skipped_rows = 0
            ambiguous_rows = 0
            unmatched_rows = 0

            for i, (file_path, rows_data) in enumerate(file_rows.items(), start=1):
                if self.stop_requested:
                    self.log("用户停止了同步。")
                    break

                file_name = os.path.basename(file_path)
                if self.progress:
                    self.progress(i - 1, total_files, f"更新: {file_name}")

                if not os.path.exists(file_path):
                    self.log(f"跳过不存在的文件: {file_path}")
                    skipped_rows += len(rows_data)
                    continue

                doc = None
                try:
                    try:
                        doc = open_word_doc_editable(word_app, file_path)
                    except Exception:
                        time.sleep(0.6)
                        doc = open_word_doc_editable(word_app, file_path)

                    if doc.Tables.Count <= 0:
                        self.log(f"跳过无表格文件: {file_path}")
                        skipped_rows += len(rows_data)
                        continue

                    table = doc.Tables(1)
                    word_key_to_rows = {}
                    word_row_cache = {}
                    word_sig_by_row = {}

                    for wr in range(2, table.Rows.Count + 1):
                        vals = []
                        for c in range(1, 14):
                            try:
                                vals.append(clean_word_text(table.Cell(wr, c).Range.Text))
                            except Exception:
                                vals.append("")
                        k = build_key(vals, key_cols)
                        if not k:
                            continue
                        word_key_to_rows.setdefault(k, []).append(wr)
                        word_row_cache[wr] = vals
                        word_sig_by_row[wr] = build_key(vals, base_cols)

                    used_word_rows = set()
                    changed = 0

                    def match_score(excel_vals, word_vals):
                        score = 0
                        for idx0 in base_cols:
                            ev = norm_key_part(excel_vals[idx0])
                            wv = norm_key_part(word_vals[idx0])
                            if not ev or not wv:
                                continue
                            if ev == wv:
                                score += 1
                        return score

                    def pick_best_row(excel_vals, primary_key, sig, used_rows):
                        candidates = word_key_to_rows.get(primary_key) or []
                        candidates = [r for r in candidates if r not in used_rows]

                        if candidates:
                            exact = [r for r in candidates if word_sig_by_row.get(r) == sig]
                            if len(exact) == 1:
                                return exact[0], "exact"
                            if len(exact) > 1:
                                return None, "ambiguous"

                            scored = []
                            for r in candidates:
                                wv = word_row_cache.get(r)
                                if not wv:
                                    continue
                                scored.append((match_score(excel_vals, wv), r))
                            scored.sort(reverse=True)
                            if not scored:
                                return None, "unmatched"
                            best_score, best_r = scored[0]
                            second_score = scored[1][0] if len(scored) > 1 else -1
                            min_needed = max(2, len(base_cols) // 3)
                            if best_score >= min_needed and best_score > second_score:
                                return best_r, "scored"
                            return None, "ambiguous"

                        scored_all = []
                        for r, wv in word_row_cache.items():
                            if r in used_rows:
                                continue
                            scored_all.append((match_score(excel_vals, wv), r))
                        scored_all.sort(reverse=True)
                        if not scored_all:
                            return None, "unmatched"
                        best_score, best_r = scored_all[0]
                        second_score = scored_all[1][0] if len(scored_all) > 1 else -1
                        min_needed = max(3, len(base_cols) // 2)
                        if best_score >= min_needed and best_score > second_score:
                            return best_r, "scored"
                        return None, "unmatched"

                    rows_data_sorted = sorted(rows_data, key=lambda x: x[0])
                    for excel_row_idx, excel_vals, primary_key, sig in rows_data_sorted:
                        target_wr, mode = pick_best_row(excel_vals, primary_key, sig, used_word_rows)
                        if target_wr is None:
                            skipped_rows += 1
                            if mode == "ambiguous":
                                ambiguous_rows += 1
                            else:
                                unmatched_rows += 1
                            continue

                        used_word_rows.add(target_wr)
                        word_vals = word_row_cache.get(target_wr, [""] * 13)

                        for col0 in update_cols:
                            new_v = fmt_excel_value(excel_vals[col0])
                            old_v = norm_key_part(word_vals[col0])
                            if new_v == old_v:
                                continue
                            try:
                                table.Cell(target_wr, col0 + 1).Range.Text = new_v
                                changed += 1
                            except Exception:
                                pass

                    if changed > 0:
                        try:
                            doc.Save()
                        except Exception as e:
                            self.log(f"保存失败 {file_path}: {e}")
                            continue
                        updated_files += 1
                        updated_cells += changed

                except Exception as e:
                    self.log(f"更新失败 {file_path}: {type(e).__name__}: {e}")
                finally:
                    if doc:
                        try:
                            doc.Close(False)
                        except Exception:
                            try:
                                doc.Close()
                            except Exception:
                                pass

            if self.progress:
                self.progress(total_files, total_files, "完成")
            self.log(f"反向同步完成：更新文件 {updated_files}/{total_files}，更新单元格 {updated_cells}，跳过记录 {skipped_rows}（无法匹配 {unmatched_rows}，匹配不唯一 {ambiguous_rows}）。")
            if updated_cells == 0:
                self.log("提示：未发生任何写入。通常是因为 Excel 行与 Word 行无法稳定匹配（字段差异/重复记录/合并单元格）。建议先保证“地点/类型/描述/发现时间”等定位字段在两边一致。")
            return True

        except Exception as e:
            self.log(f"Word服务异常: {type(e).__name__}: {e}")
            return False
        finally:
            if word_app:
                try:
                    word_app.Quit()
                except Exception:
                    pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def process_source(self, source_path, target_excel, overwrite=False, incremental=False):
        try:
            pythoncom.CoInitialize()
        except Exception:
            pass

        self.log(f"开始处理: {source_path}")
        
        if not os.path.exists(target_excel):
            self.log(f"错误: 找不到目标Excel文件: {target_excel}")
            return False

        # 1. Collect all Word files
        word_files = []
        if os.path.isfile(source_path):
             if source_path.lower().endswith(('.doc', '.docx')) and not os.path.basename(source_path).startswith('~$'):
                word_files.append(source_path)
        elif os.path.isdir(source_path):
            self.log(f"正在扫描文件夹: {source_path}")
            for root, dirs, files in os.walk(source_path):
                for file in files:
                    if file.lower().endswith(('.doc', '.docx')) and not file.startswith('~$'):
                        word_files.append(os.path.join(root, file))
        else:
            self.log(f"错误: 找不到源文件或文件夹: {source_path}")
            return False

        try:
            word_files.sort()
        except Exception:
            pass

        if incremental:
            try:
                self._normalize_excel_rows(target_excel)
            except PermissionError:
                self.log("错误: 目标Excel文件被占用 (Permission denied)。")
                messagebox.showwarning("文件被占用", "无法写入目标Excel文件。\n\n请检查该文件是否在Excel中打开。\n请关闭文件后再次点击“导入并同步”。")
                return False

            processed = self._load_processed_paths_from_excel(target_excel)
            if processed:
                # 1. Handle deleted files
                current_files_set = {os.path.normcase(os.path.normpath(p)) for p in word_files}
                deleted_files = processed - current_files_set
                
                if deleted_files:
                    self.log(f"发现 {len(deleted_files)} 个历史文件已被删除，正在同步清理Excel记录...")
                    removed_count = self._remove_rows_by_paths(target_excel, deleted_files)
                    self.log(f"已清理 {removed_count} 条无效记录。")
                    try:
                        self._normalize_excel_rows(target_excel)
                    except PermissionError:
                        self.log("错误: 目标Excel文件被占用 (Permission denied)。")
                        messagebox.showwarning("文件被占用", "无法写入目标Excel文件。\n\n请检查该文件是否在Excel中打开。\n请关闭文件后再次点击“导入并同步”。")
                        return False

                # 2. Handle new files
                before = len(word_files)
                word_files = [p for p in word_files if os.path.normcase(os.path.normpath(p)) not in processed]
                
                if not word_files:
                    if deleted_files:
                        self.log("未发现新Word文档，同步完成。")
                    else:
                        self.log("未发现新Word文档，无需同步。")
                    
                    if self.progress:
                        self.progress(before, before, "完成")
                    return True
            else:
                self.log("提示: 未能从Excel读取历史路径，将执行全量同步。")

        total_files = len(word_files)
        self.log(f"共发现 {total_files} 个Word文件。")
        
        if self.progress:
            self.progress(0, total_files, "准备开始...")

        if not word_files:
            return True

        # 2. Extract data
        extracted_rows = []

        word = None
        temp_dir_obj = None
        try:
            pythoncom.CoInitialize()
            temp_dir_obj = tempfile.TemporaryDirectory()
            temp_dir = temp_dir_obj.name

            def close_word(app):
                if not app:
                    return
                try:
                    app.Quit()
                except Exception:
                    pass

            def kill_all_winword():
                try:
                    subprocess.run(
                        ["taskkill", "/F", "/IM", "WINWORD.EXE"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
                    )
                except Exception:
                    pass

            def create_word():
                kill_all_winword()
                app = None
                try:
                    app = win32com.client.DispatchEx("Word.Application")
                except Exception as e:
                    # Try WPS fallback
                    try:
                        app = win32com.client.DispatchEx("Kwps.Application")
                    except Exception:
                        try:
                            app = win32com.client.DispatchEx("Wps.Application")
                        except Exception:
                            raise e

                try:
                    app.Visible = False
                except Exception:
                    pass
                try:
                    app.DisplayAlerts = 0
                except Exception:
                    pass
                try:
                    app.AutomationSecurity = 3
                except Exception:
                    pass
                return app

            def is_rpc_unavailable(err):
                try:
                    args = getattr(err, "args", None)
                    if args and len(args) >= 1 and int(args[0]) == -2147023174:
                        return True
                except Exception:
                    pass
                return False

            def ensure_word_alive(app):
                if not app:
                    return False
                try:
                    _ = app.Version
                    return True
                except Exception:
                    return False

            word = create_word()
            time.sleep(0.8)
            consecutive_rpc_failures = 0

            for i, file_path in enumerate(word_files):
                while getattr(self, "paused", False):
                    if self.stop_requested:
                        break
                    time.sleep(0.1)

                if self.stop_requested:
                    self.log("用户停止了操作。")
                    break

                file_name = os.path.basename(file_path)
                self.log(f"正在读取 ({i+1}/{total_files}): {file_name}")
                if self.progress:
                    self.progress(i + 1, total_files, f"读取: {file_name}")

                success = False
                last_error = None

                for attempt in range(3):
                    doc = None
                    try:
                        if not ensure_word_alive(word):
                            close_word(word)
                            word = None
                            word = create_word()
                            time.sleep(0.8)

                        try:
                            doc = self._open_word_doc(word, file_path)
                        except Exception as e:
                            tmp_name = f"{i+1:04d}_{self._safe_temp_name(file_name)}"
                            tmp_path = os.path.join(temp_dir, tmp_name)
                            shutil.copy2(file_path, tmp_path)
                            doc = self._open_word_doc(word, tmp_path)

                        if doc.Tables.Count > 0:
                            table = doc.Tables(1)
                            row_count = table.Rows.Count
                            if row_count > 1:
                                for r in range(2, row_count + 1):
                                    row_data = []
                                    try:
                                        for c in range(1, 14):
                                            try:
                                                cell_text = table.Cell(r, c).Range.Text
                                                cell_text = cell_text.replace('\r', '').replace('\x07', '').strip()
                                                row_data.append(cell_text)
                                            except Exception:
                                                row_data.append("")

                                        has_content = False
                                        if len(row_data) > 1:
                                            for cell in row_data[1:]:
                                                if cell.strip():
                                                    has_content = True
                                                    break

                                        if has_content:
                                            row_data.append(file_path)
                                            extracted_rows.append(row_data)
                                    except Exception:
                                        pass
                            else:
                                self.log(f"  警告: {file_name} 表格行数不足")
                        else:
                            self.log(f"  警告: {file_name} 中没有表格")

                        try:
                            doc.Close(False)
                        except Exception:
                            pass
                        doc = None
                        success = True
                        break
                    except Exception as e:
                        last_error = e
                        try:
                            if doc:
                                doc.Close(False)
                        except Exception:
                            pass
                        if is_rpc_unavailable(e):
                            close_word(word)
                            word = None
                            kill_all_winword()
                            time.sleep(1.2)
                            try:
                                word = create_word()
                                time.sleep(0.8)
                            except Exception:
                                word = None
                            consecutive_rpc_failures += 1
                        else:
                            consecutive_rpc_failures = 0
                        time.sleep(0.8)

                if not success:
                    if last_error is None:
                        self.log(f"  错误: 无法读取文件 {file_name}")
                    else:
                        if isinstance(last_error, Exception) and "无效的类字符串" in str(last_error):
                             self.log(f"  错误: 无法启动 Word 或 WPS。请确认已安装 Microsoft Office 或 WPS Office。")
                        else:
                             self.log(f"  错误: 无法读取文件 {file_name}（{type(last_error).__name__}: {last_error}）")

                    if is_rpc_unavailable(last_error) or consecutive_rpc_failures >= 2:
                        isolated_word = None
                        isolated_doc = None
                        isolated_error = None
                        try:
                            isolated_word = create_word()
                            time.sleep(0.8)

                            try:
                                isolated_doc = self._open_word_doc(isolated_word, file_path)
                            except Exception:
                                tmp_name = f"isolated_{i+1:04d}_{self._safe_temp_name(file_name)}"
                                tmp_path = os.path.join(temp_dir, tmp_name)
                                shutil.copy2(file_path, tmp_path)
                                isolated_doc = self._open_word_doc(isolated_word, tmp_path)

                            if isolated_doc.Tables.Count > 0:
                                table = isolated_doc.Tables(1)
                                row_count = table.Rows.Count
                                if row_count > 1:
                                    for r in range(2, row_count + 1):
                                        row_data = []
                                        try:
                                            for c in range(1, 14):
                                                try:
                                                    cell_text = table.Cell(r, c).Range.Text
                                                    cell_text = cell_text.replace('\r', '').replace('\x07', '').strip()
                                                    row_data.append(cell_text)
                                                except Exception:
                                                    row_data.append("")

                                            has_content = False
                                            if len(row_data) > 1:
                                                for cell in row_data[1:]:
                                                    if cell.strip():
                                                        has_content = True
                                                        break
                                            if has_content:
                                                row_data.append(file_path)
                                                extracted_rows.append(row_data)
                                        except Exception:
                                            pass
                                    self.log(f"  修复: 已通过隔离模式读取 {file_name}")
                                else:
                                    self.log(f"  警告: {file_name} 表格行数不足")
                            else:
                                self.log(f"  警告: {file_name} 中没有表格")
                        except Exception as e2:
                            isolated_error = e2
                        finally:
                            try:
                                if isolated_doc:
                                    isolated_doc.Close(False)
                            except Exception:
                                pass
                            try:
                                if isolated_word:
                                    isolated_word.Quit()
                            except Exception:
                                pass

                        if isolated_error is not None:
                            self.log(f"  错误: 隔离模式仍失败 {file_name}（{type(isolated_error).__name__}: {isolated_error}）")
        except Exception as e:
            self.log(f"错误: 读取Word时发生异常（{type(e).__name__}: {e}）")
            return False
        finally:
            try:
                close_word(word)
            except Exception:
                pass
            try:
                kill_all_winword()
            except Exception:
                pass
            try:
                if temp_dir_obj:
                    temp_dir_obj.cleanup()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

        if self.stop_requested:
            return False

        if not extracted_rows:
            self.log("未提取到任何数据。")
            if self.progress: self.progress(total_files, total_files, "完成")
            return True

        self.log(f"提取完成，共 {len(extracted_rows)} 条记录。正在写入Excel...")
        if self.progress:
            self.progress(total_files, total_files, "正在写入Excel...")

        # 3. Write to Excel
        try:
            wrote = self._write_rows_to_excel(target_excel, extracted_rows, overwrite=overwrite)
            try:
                self._normalize_excel_rows(target_excel)
            except PermissionError:
                self.log("错误: 目标Excel文件被占用 (Permission denied)。")
                messagebox.showwarning("文件被占用", "无法写入目标Excel文件。\n\n请检查该文件是否在Excel中打开。\n请关闭文件后再次点击“开始处理”。")
                return False
            if overwrite:
                self.log(f"写入成功！已刷新 {wrote} 条记录，已保存到: {target_excel}")
            else:
                self.log(f"写入成功！新增 {wrote} 条记录，已保存到: {target_excel}")
            return True

        except PermissionError:
            self.log("错误: 目标Excel文件被占用 (Permission denied)。")
            messagebox.showwarning("文件被占用", "无法写入目标Excel文件。\n\n请检查该文件是否在Excel中打开。\n请关闭文件后再次点击“开始处理”。")
            return False
        except Exception as e:
            self.log(f"写入Excel失败: {e}")
            return False

class StatisticsPanel(ttk.Frame):
    def __init__(self, parent, excel_path, app_instance=None):
        super().__init__(parent)
        self.excel_path = excel_path
        self.app = app_instance
        self.df = None
        self._loaded_path = None
        self._loaded_mtime = None
        self._resize_job = None
        self._last_canvas_size = None
        self._redraw_job = None
        self._redraw_attempts = 0
        self._redraw_stable = 0
        self._redraw_last = None
        self._layout_mode = None
        self.file_path_map = {}
        
        # List View State
        self.list_data_source = None
        self.sort_col = None
        self.sort_reverse = False
        
        self.setup_ui()

    def setup_ui(self):
        # --- Top Control Bar (Apple-style Layout) ---
        # Main Container with increased padding and white background
        control_bar = ttk.Frame(self, style="Card.TFrame", padding=(20, 15, 20, 15))
        control_bar.pack(fill=X, pady=(0, 1)) # Small gap below

        # --- Left: Data Actions ---
        action_group = ttk.Frame(control_bar, style="Card.TFrame")
        action_group.pack(side=LEFT)
        
        self.btn_load = ttk.Button(action_group, text=" 同步数据", command=self.load_data, bootstyle="primary", width=10)
        self.btn_load.pack(side=LEFT, padx=(0, 10))

        if self.app:
            self.btn_sync = ttk.Button(action_group, text=" 导入并同步", command=self.on_sync, bootstyle="success", width=12)
            self.btn_sync.pack(side=LEFT, padx=(0, 10))
        
        # Divider
        ttk.Separator(control_bar, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=20, pady=5)

        # --- Center Left: Filters ---
        filter_group = ttk.Frame(control_bar, style="Card.TFrame")
        filter_group.pack(side=LEFT)

        def create_filter(parent, label, variable, values, width, command=None):
            f_box = ttk.Frame(parent, style="Card.TFrame")
            f_box.pack(side=LEFT, padx=(0, 15))
            
            lbl = ttk.Label(f_box, text=label, font=("Microsoft YaHei UI", 9), foreground="#666666", background="#FFFFFF")
            lbl.pack(side=LEFT, padx=(0, 8))
            
            cb = ttk.Combobox(f_box, textvariable=variable, values=values, width=width, state="readonly", bootstyle="default")
            cb.pack(side=LEFT)
            if command:
                cb.bind("<<ComboboxSelected>>", command)
            return cb

        self.year_var = tk.StringVar(value="全部")
        self.year_cb = create_filter(filter_group, "年份", self.year_var, ["全部"], 8, self.apply_filter)
        
        self.month_var = tk.StringVar(value="全部")
        months = ["全部"] + [f"{i}月" for i in range(1, 13)]
        self.month_cb = create_filter(filter_group, "月份", self.month_var, months, 6, self.apply_filter)
        
        self.status_filter_var = tk.StringVar(value="全部状态")
        self.cb_status = create_filter(filter_group, "状态", self.status_filter_var, ["全部状态", "未销号", "已销号"], 10, lambda e: self.refresh_tree_view())

        # --- Right: Search & Tools ---
        right_group = ttk.Frame(control_bar, style="Card.TFrame")
        right_group.pack(side=RIGHT)

        # Export (Far Right)
        self.btn_export = ttk.Button(right_group, text=" 导出图表", command=self.export_chart, bootstyle="info-outline")
        self.btn_export.pack(side=RIGHT, padx=(15, 0))

        # Search Box Area
        search_box = ttk.Frame(right_group, style="Card.TFrame")
        search_box.pack(side=RIGHT)
        
        self.search_var = tk.StringVar()
        self.entry_search = ttk.Entry(search_box, textvariable=self.search_var, width=20, bootstyle="secondary")
        self.entry_search.pack(side=LEFT, padx=(0, 5))
        self.entry_search.bind("<Return>", lambda e: self.refresh_tree_view())
        
        ttk.Button(search_box, text="查询", command=self.refresh_tree_view, bootstyle="secondary-outline", width=6).pack(side=LEFT, padx=(0, 5))
        ttk.Button(search_box, text="重置", command=self.reset_list_filters, bootstyle="link-secondary").pack(side=LEFT)

        # Status Info (Flexible Spacer)
        self.lbl_status = ttk.Label(control_bar, text="请先同步数据", bootstyle="secondary", background="#FFFFFF", font=("Microsoft YaHei UI", 8))
        self.lbl_status.pack(side=LEFT, padx=30)

        # --- Content Area ---
        self.content_area = ttk.Frame(self)
        self.content_area.pack(fill=BOTH, expand=YES, pady=0)
        
        # View 1: Dashboard
        self.view_dashboard = ttk.Frame(self.content_area, padding=10) # Add padding for dashboard
        self.setup_dashboard_tab(self.view_dashboard)
        
        # View 2: Detail List
        self.view_details = ttk.Frame(self.content_area, padding=0)
        self.setup_details_tab(self.view_details)
        
        # Default View
        self.current_view = None
        self.switch_view("chart")

    def switch_view(self, view_name):
        if self.current_view:
            self.current_view.pack_forget()
            
        if view_name == "chart":
            self.view_dashboard.pack(fill=BOTH, expand=YES)
            self.current_view = self.view_dashboard
            self.request_redraw()
        elif view_name == "list":
            self.view_details.pack(fill=BOTH, expand=YES)
            self.current_view = self.view_details

    def setup_dashboard_tab(self, parent):
        # Summary Cards (Top)
        self.cards_frame = ttk.Frame(parent)
        self.cards_frame.pack(fill=X, pady=5)
        
        self.card_total = self.create_card(self.cards_frame, "缺陷总数", "0", "info")
        self.card_open = self.create_card(self.cards_frame, "未销号", "0", "danger")
        self.card_closed = self.create_card(self.cards_frame, "已销号", "0", "success")
        
        # Charts Area (Middle)
        self.charts_frame = ttk.Frame(parent)
        self.charts_frame.pack(fill=BOTH, expand=YES, pady=5)
        
        self.fig = Figure(figsize=(10, 5), dpi=100, constrained_layout=True)
        self.fig.patch.set_facecolor('#F8F9FA') # Match light theme bg roughly
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.charts_frame)
        self.canvas_widget = self.canvas.get_tk_widget()
        self.canvas_widget.pack(fill=BOTH, expand=YES)
        self.canvas_widget.bind("<Configure>", self.on_resize, add="+")
        self.after(0, self._sync_figure_dpi_to_tk)

    def setup_details_tab(self, parent):
        # --- Treeview ---
        # Increase row height for better readability
        style = ttk.Style()
        for style_name in ("Treeview", "primary.Treeview"):
            try:
                style.configure(style_name, rowheight=40)
            except Exception:
                pass

        columns = ("serial", "discovery_date", "location", "type", "status", "date", "action")
        self.tree = ttk.Treeview(parent, columns=columns, show="headings", bootstyle="primary")
        
        self.tree.heading("serial", text="序号", command=lambda: self.on_sort_column("serial"))
        self.tree.heading("discovery_date", text="缺陷发现时间", command=lambda: self.on_sort_column("discovery_date"))
        self.tree.heading("location", text="设备缺陷地点", command=lambda: self.on_sort_column("location"))
        self.tree.heading("type", text="设备缺陷类型", command=lambda: self.on_sort_column("type"))
        self.tree.heading("status", text="状态", command=lambda: self.on_sort_column("status"))
        self.tree.heading("date", text="销号时间", command=lambda: self.on_sort_column("date"))
        self.tree.heading("action", text="操作")
        
        self.tree.column("serial", width=60, anchor="center")
        self.tree.column("discovery_date", width=150, anchor="center")
        self.tree.column("location", width=250, anchor="center")
        self.tree.column("type", width=150, anchor="center")
        self.tree.column("status", width=100, anchor="center")
        self.tree.column("date", width=150, anchor="center")
        self.tree.column("action", width=100, anchor="center")
        
        # Scrollbar
        vsb = ttk.Scrollbar(parent, orient="vertical", command=self.tree.yview, bootstyle="round")
        self.tree.configure(yscrollcommand=vsb.set)
        
        self.tree.pack(side=LEFT, fill=BOTH, expand=YES)
        vsb.pack(side=RIGHT, fill=Y)
        
        self.tree.bind("<Double-1>", self.on_tree_double_click)

        # Optimize Mouse Wheel Scrolling
        def _on_mousewheel(event):
            try:
                # Windows: event.delta is usually 120/-120
                # Accelerate scrolling speed (factor of 3)
                delta = int(-1 * (event.delta / 120) * 3)
                self.tree.yview_scroll(delta, "units")
            except Exception:
                pass
            return "break"

        self.tree.bind("<MouseWheel>", _on_mousewheel)
        
        # Tooltip or instructions removed as per user request


    def on_sync(self):
        if self.app:
            self.app.run_sync_process_from_stats()

    def on_resize(self, event):
        self._last_canvas_size = (event.width, event.height)
        if self._resize_job is not None:
            try:
                self.after_cancel(self._resize_job)
            except Exception:
                pass
        self._resize_job = self.after(120, self._on_resize_debounced)

    def _get_tk_dpi(self):
        try:
            dpi = float(self.canvas_widget.winfo_fpixels('1i'))
            if dpi > 0:
                return dpi
        except Exception:
            pass
        return float(self.fig.dpi)

    def _sync_figure_dpi_to_tk(self):
        try:
            dpi = float(self._get_tk_dpi())
        except Exception:
            return
        if dpi <= 0:
            return
        try:
            current = float(self.fig.dpi)
        except Exception:
            current = dpi
        if abs(current - dpi) < 0.5:
            return
        try:
            self.fig.set_dpi(dpi)
        except Exception:
            return

    def _layout_mode_for_width(self, width):
        try:
            w = int(width)
        except Exception:
            w = 0
        if w > 1 and w < 800:
            return "vertical"
        return "horizontal"

    def _on_resize_debounced(self):
        self._resize_job = None
        self._sync_figure_dpi_to_tk()
        try:
            w = int(self.canvas_widget.winfo_width())
        except Exception:
            w = 0
        if w <= 1 and self._last_canvas_size:
            w = int(self._last_canvas_size[0])
        new_mode = self._layout_mode_for_width(w)
        if self.df is not None and new_mode != self._layout_mode:
            self.render_charts()
            return
        try:
            self.canvas.draw_idle()
        except Exception:
            pass

    def request_redraw(self):
        if self._redraw_job is not None:
            try:
                self.after_cancel(self._redraw_job)
            except Exception:
                pass
        self._redraw_attempts = 0
        self._redraw_stable = 0
        self._redraw_last = None
        self._redraw_job = self.after(0, self._redraw_tick)

    def _redraw_tick(self):
        self._redraw_job = None
        self._redraw_attempts += 1
        try:
            self.update_idletasks()
            w = int(self.canvas_widget.winfo_width())
            h = int(self.canvas_widget.winfo_height())
        except Exception:
            return

        if w < 60 or h < 60:
            if self._redraw_attempts < 15:
                self._redraw_job = self.after(60, self._redraw_tick)
            return

        if self._redraw_last == (w, h):
            self._redraw_stable += 1
        else:
            self._redraw_last = (w, h)
            self._redraw_stable = 0

        if self._redraw_stable >= 2 or self._redraw_attempts >= 15:
            try:
                self._sync_figure_dpi_to_tk()
                new_mode = self._layout_mode_for_width(w)
                if self.df is not None and new_mode != self._layout_mode:
                    self.render_charts()
                    return
                self.canvas.draw_idle()
            except Exception:
                return
            return

        self._redraw_job = self.after(80, self._redraw_tick)

    def create_card(self, parent, title, value, bootstyle="primary"):
        frame = ttk.Frame(parent, bootstyle="light", padding=10)
        # Use expand=YES to distribute width evenly
        frame.pack(side=LEFT, fill=BOTH, expand=YES, padx=5)
        
        # Use a localized style for the card content
        ttk.Label(frame, text=title, font=("Microsoft YaHei UI", 10), bootstyle="secondary").pack(anchor=W)
        lbl = ttk.Label(frame, text=value, font=("Microsoft YaHei UI", 24, "bold"), bootstyle=bootstyle)
        lbl.pack(pady=5)
        return lbl

    def load_data(self, force=False, silent=False):
        path = self.excel_path.get()
        if not os.path.exists(path):
            if not silent:
                messagebox.showerror("错误", "找不到Excel文件")
            return

        try:
            try:
                mtime = os.path.getmtime(path)
            except Exception:
                mtime = None
            if not force and self.df is not None and self._loaded_path == path and self._loaded_mtime == mtime:
                self.lbl_status.config(text=f"数据已就绪: {time.strftime('%H:%M:%S')}")
                self.request_redraw()
                if not silent:
                    try:
                        winsound.MessageBeep()
                    except Exception:
                        pass
                    messagebox.showinfo("提示", "数据无需更新，已是最新状态。")
                return

            df = pd.read_excel(path, header=2)
            required_cols = ['设备缺陷类型', '销号时间', '设备缺陷地点']
            if not all(col in df.columns for col in required_cols):
                df = pd.read_excel(path) 
            
            filter_cols = ['设备缺陷地点', '设备缺陷类型', '设备缺陷描述']
            valid_cols = [c for c in filter_cols if c in df.columns]
            if valid_cols:
                df = df.dropna(subset=valid_cols, how='all')
            
            self.df = df
            self._loaded_path = path
            self._loaded_mtime = mtime
            
            if "销号时间" in self.df.columns:
                self.df["销号时间"] = self._parse_datetime_series(self.df["销号时间"])
            self._refresh_year_options(self.df)
            
            self.update_dashboard(self.df)
            self.lbl_status.config(text=f"数据已更新: {time.strftime('%H:%M:%S')}")
            self.btn_export.config(state="normal")
            self.request_redraw()
            
            if not silent:
                try:
                    winsound.MessageBeep()
                except Exception:
                    pass
                messagebox.showinfo("提示", "数据加载成功！")
            
        except PermissionError:
            if not silent:
                messagebox.showwarning("提示", "请关闭Excel文件后再读取")
        except Exception as e:
            if not silent:
                messagebox.showerror("错误", str(e))

    def apply_filter(self, event=None):
        if self.df is None:
            return

        filtered_df = self.filter_dataframe(
            self.df,
            apply_date_filters=True,
            apply_status_filter=False,
            apply_search_filter=False,
        )
        self.update_dashboard(filtered_df)
        self.request_redraw()

    def _get_closed_mask(self, df):
        if df is None or df.empty:
            return pd.Series([], dtype=bool)
        if '销号时间' not in df.columns:
            return pd.Series([False] * len(df), index=df.index)
        try:
            return pd.to_datetime(df['销号时间'], errors='coerce').notna()
        except Exception:
            return df['销号时间'].notna()

    def _get_date_candidate_columns(self, df):
        if df is None or df.empty or not hasattr(df, "columns"):
            return []
        cols = []
        for c in df.columns:
            if c == "销号时间":
                continue
            s = str(c)
            if "时间" in s or "日期" in s:
                cols.append(c)
        return cols

    def _parse_datetime_series(self, series):
        try:
            return pd.to_datetime(series, errors="coerce", format="mixed")
        except TypeError:
            pass
        try:
            dt = pd.to_datetime(series, errors="coerce")
        except Exception:
            try:
                dt = pd.to_datetime(series.astype(str), errors="coerce")
            except Exception:
                return pd.Series([pd.NaT] * len(series), index=series.index)
        try:
            if dt.notna().all():
                return dt
        except Exception:
            return dt

        try:
            s = series.astype(str)
        except Exception:
            return dt

        try:
            s = s.str.strip()
            s = s.str.replace("年", "-", regex=False)
            s = s.str.replace("月", "-", regex=False)
            s = s.str.replace("日", "", regex=False)
            s = s.str.replace(".", "-", regex=False)
            s = s.str.replace("/", "-", regex=False)
            s = s.str.replace(r"\s+", " ", regex=True).str.strip()
            dt2 = pd.to_datetime(s, errors="coerce")
            return dt.where(dt.notna(), dt2)
        except Exception:
            return dt

    def _date_column_priority_key(self, col, non_null_count=0):
        s = str(col)
        keywords = ["发现", "发生", "缺陷", "登记", "填报", "上报", "录入", "创建"]
        kw_idx = next((i for i, kw in enumerate(keywords) if kw in s), len(keywords))
        exact = [
            "缺陷时间", "缺陷日期",
            "发生时间", "发生日期",
            "发现时间", "发现日期",
            "登记时间", "登记日期",
            "填报时间", "填报日期",
            "上报时间", "上报日期",
            "录入时间", "录入日期",
            "创建时间", "创建日期",
            "日期",
        ]
        exact_idx = next((i for i, name in enumerate(exact) if name == s), len(exact))
        return (-int(non_null_count), exact_idx, kw_idx, s)

    def _refresh_year_options(self, df):
        filter_dt = self._get_filter_datetime(df)
        years = (
            filter_dt.dt.year.dropna()
            .unique()
            .astype(int)
            .tolist()
        )
        years = sorted(set(years))
        values = ["全部"] + [str(y) for y in years]
        self.year_cb["values"] = values
        current = (self.year_var.get() or "").strip()
        if not current or current not in values:
            self.year_var.set("全部")

    def _choose_reference_date_column(self, df):
        if df is None or df.empty:
            return None
        if not hasattr(df, "columns"):
            return None

        candidates = self._get_date_candidate_columns(df)
        if not candidates:
            return None

        best_col = None
        best_key = None
        for c in candidates:
            dt = self._parse_datetime_series(df[c])
            cnt = int(dt.notna().sum())
            key = self._date_column_priority_key(c, cnt)
            if best_key is None or key < best_key:
                best_key = key
                best_col = c
        return best_col

    def _get_filter_datetime(self, df):
        if df is None or df.empty:
            return pd.Series([], dtype="datetime64[ns]")

        close_dt = None
        if "销号时间" in df.columns:
            try:
                close_dt = self._parse_datetime_series(df["销号时间"])
            except Exception:
                close_dt = None

        ref_dt = None
        candidates = self._get_date_candidate_columns(df)
        if candidates:
            parsed = []
            for c in candidates:
                dt = self._parse_datetime_series(df[c])
                parsed.append((c, dt, int(dt.notna().sum())))
            parsed.sort(key=lambda x: self._date_column_priority_key(x[0], x[2]))
            for _, dt, _ in parsed:
                if ref_dt is None:
                    ref_dt = dt
                else:
                    ref_dt = ref_dt.where(ref_dt.notna(), dt)

        if close_dt is None and ref_dt is None:
            return pd.Series([pd.NaT] * len(df), index=df.index)
        if close_dt is None:
            return ref_dt
        if ref_dt is None:
            return close_dt
        # Prioritize discovery date (ref_dt) for filtering, fallback to close_dt
        return ref_dt.where(ref_dt.notna(), close_dt)

    def filter_dataframe(
        self,
        df,
        apply_date_filters=True,
        apply_status_filter=True,
        apply_search_filter=True,
    ):
        if df is None or df.empty:
            return df

        out = df.copy()
        closed_mask = self._get_closed_mask(out)
        filter_dt = self._get_filter_datetime(out)

        if apply_date_filters:
            year = (self.year_var.get() or "").strip()
            month_str = (self.month_var.get() or "").strip()

            if year and year != "全部":
                try:
                    y = int(year)
                    out = out[filter_dt.dt.year == y]
                    closed_mask = closed_mask.loc[out.index]
                    filter_dt = filter_dt.loc[out.index]
                except Exception:
                    pass

            if month_str and month_str != "全部":
                try:
                    m = int(month_str.replace("月", ""))
                    out = out[filter_dt.dt.month == m]
                    closed_mask = closed_mask.loc[out.index]
                except Exception:
                    pass

        if apply_status_filter and not out.empty:
            status = (self.status_filter_var.get() or "").strip()
            if status == "未销号":
                out = out[~closed_mask.loc[out.index]]
            elif status == "已销号":
                out = out[closed_mask.loc[out.index]]

        if apply_search_filter and not out.empty:
            query = (self.search_var.get() or "").strip()
            if query:
                searchable_cols = [c for c in ["序号", "设备缺陷地点", "设备缺陷类型", "销号时间"] if c in out.columns]
                if searchable_cols:
                    mask = pd.Series(False, index=out.index)
                    for c in searchable_cols:
                        try:
                            mask = mask | out[c].astype(str).str.contains(query, case=False, na=False, regex=False)
                        except Exception:
                            pass
                    out = out[mask]

        return out

    def update_dashboard(self, df):
        # 1. Update Cards
        total = len(df)
        closed = df['销号时间'].notna().sum()
        pending = total - closed
        
        self.card_total.config(text=str(total))
        self.card_open.config(text=str(pending))
        self.card_closed.config(text=str(closed))
        
        # 2. Update Charts
        self.render_charts(df)
        
        # 3. Update Detail List
        self.update_detail_list(df)

    def update_detail_list(self, df):
        self.list_data_source = df
        self.refresh_tree_view()

    def reset_list_filters(self):
        self.search_var.set("")
        self.status_filter_var.set("全部状态")
        self.sort_col = None
        self.sort_reverse = False
        # Reset headers
        for c in ["serial", "discovery_date", "location", "type", "status", "date"]:
            self.tree.heading(c, text=self.tree.heading(c, "text").replace(" ▲", "").replace(" ▼", ""))
        self.refresh_tree_view()

    def on_sort_column(self, col):
        if self.sort_col == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_col = col
            self.sort_reverse = False
            
        # Update heading indicators
        for c in ["serial", "discovery_date", "location", "type", "status", "date"]:
            text = self.tree.heading(c, "text").replace(" ▲", "").replace(" ▼", "")
            if c == self.sort_col:
                text += " ▼" if self.sort_reverse else " ▲"
            self.tree.heading(c, text=text)
            
        self.refresh_tree_view()

    def refresh_tree_view(self):
        # Clear existing
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.file_path_map.clear()
        
        if self.list_data_source is None or self.list_data_source.empty:
            return
            
        df = self.list_data_source.copy()

        df = self.filter_dataframe(
            df,
            apply_date_filters=False,
            apply_status_filter=True,
            apply_search_filter=True,
        )
        if df is None or df.empty:
            return
            
        # 2. Apply Sorting
        discovery_col = self._choose_reference_date_column(df)

        if self.sort_col and not df.empty:
            col_map = {
                "serial": "序号",
                "location": "设备缺陷地点",
                "type": "设备缺陷类型",
                "discovery_date": discovery_col,
                "status": "销号时间",
                "date": "销号时间"
            }
            
            df_col = col_map.get(self.sort_col)
            if df_col:
                ascending = not self.sort_reverse
                if self.sort_col == "serial":
                    if '序号' in df.columns:
                        try:
                            df['__serial_sort'] = pd.to_numeric(df['序号'], errors='coerce')
                            df = df.sort_values(by='__serial_sort', ascending=ascending)
                        except:
                            df = df.sort_values(by='序号', ascending=ascending)
                    else:
                        df = df.sort_index(ascending=ascending)
                elif self.sort_col == "status":
                     df['__is_closed'] = df['销号时间'].notna()
                     df = df.sort_values(by='__is_closed', ascending=ascending)
                else:
                    df = df.sort_values(by=df_col, ascending=ascending, na_position='last')

        for index, row in df.iterrows():
            serial = ""
            try:
                if '序号' in df.columns:
                    serial = row.get('序号', "")
                elif len(row) > 0:
                    serial = row.iloc[0]
            except Exception:
                serial = ""
            if serial is None or str(serial).strip() == "" or str(serial).strip().lower() == "nan":
                serial = index + 1
            loc = row.get('设备缺陷地点', '')
            dtype = row.get('设备缺陷类型', '')
            
            discovery_str = "-"
            if discovery_col and discovery_col in row:
                try:
                    d_val = row.get(discovery_col)
                    d_ts = pd.to_datetime(d_val, errors='coerce')
                    if pd.notna(d_ts):
                        discovery_str = d_ts.strftime('%Y-%m-%d')
                except Exception:
                    pass

            date_val = row.get('销号时间')
            date_ts = pd.to_datetime(date_val, errors='coerce')
            is_closed = pd.notna(date_ts)
            status_text = "✅ 已销号" if is_closed else "🔴 未销号"
            date_str = date_ts.strftime('%Y-%m-%d') if is_closed else "-"
            
            source_path = ""
            try:
                for v in reversed(list(row.values)):
                    if not isinstance(v, str):
                        continue
                    s = v.strip()
                    if not s:
                        continue
                    if (":\\" in s or "\\" in s or "/" in s) and s.lower().endswith((".doc", ".docx")):
                        source_path = s
                        break
            except Exception:
                source_path = ""
            
            # Insert into Treeview
            item_id = self.tree.insert("", "end", values=(serial, discovery_str, loc, dtype, status_text, date_str, "📂 打开"))
            
            # Store file path
            if source_path and os.path.exists(source_path):
                self.file_path_map[item_id] = source_path
            
            if not is_closed:
                self.tree.item(item_id, tags=("open",))
            else:
                self.tree.item(item_id, tags=("closed",))

        self.tree.tag_configure("open", foreground="red")
        self.tree.tag_configure("closed", foreground="green")

    def on_tree_double_click(self, event):
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return
            
        path = self.file_path_map.get(item_id)
        if path and os.path.exists(path):
            self._monitor_word_file(path)
        else:
            messagebox.showwarning("提示", "该条记录未关联到Word文档路径（可能是历史数据）。\n建议点击“同步数据”后再试。")

    def _monitor_word_file(self, path):
        def task():
            try:
                pythoncom.CoInitialize()
                word = None
                try:
                    word = win32com.client.Dispatch("Word.Application")
                except Exception:
                    try:
                        word = win32com.client.Dispatch("Kwps.Application")
                    except Exception:
                        word = win32com.client.Dispatch("Wps.Application")
                
                word.Visible = True
                try:
                    word.WindowState = 1  # wdWindowStateMaximize
                    word.Activate()
                except Exception:
                    pass
                doc = word.Documents.Open(path)
                doc_name = doc.Name
                
                while True:
                    time.sleep(1)
                    try:
                        found = False
                        for d in word.Documents:
                            if d.Name == doc_name:
                                found = True
                                break
                        if not found:
                            break
                    except Exception:
                        break
                
                if self.app and hasattr(self.app, 'processor'):
                     self.app.log_message(f"检测到文档关闭: {os.path.basename(path)}，正在更新数据...")
                     target_excel = self.excel_path.get()
                     success = self.app.processor.update_single_file(path, target_excel)
                     if success:
                         self.root.after(0, lambda: self.load_data(force=True, silent=True))
                         self.root.after(0, lambda: messagebox.showinfo("自动同步", f"文档 {os.path.basename(path)} 已更新并同步！"))
                     else:
                         self.root.after(0, lambda: messagebox.showerror("自动同步失败", "无法更新数据，请检查日志。"))

            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("错误", f"打开文件失败: {e}"))
                
        threading.Thread(target=task, daemon=True).start()

    def render_charts(self, df=None):
        if df is None:
            df = self.df
        if df is None:
            return

        try:
            self.update_idletasks()
        except Exception:
            pass

        self.fig.clear()
        self.fig.set_constrained_layout(True)
        
        # Check Theme for Colors
        current_theme = ttk.Style().theme_use()
        is_dark = "dark" in current_theme
        text_color = '#FFFFFF' if is_dark else '#333333'
        bg_color = '#222222' if is_dark else '#F8F9FA'
        grid_color = '#555555' if is_dark else '#DDDDDD'
        
        self.fig.patch.set_facecolor(bg_color)
        
        # Set Global Font & Colors
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'Arial Unicode MS']
        plt.rcParams['axes.unicode_minus'] = False
        plt.rcParams['text.color'] = text_color
        plt.rcParams['axes.labelcolor'] = text_color
        plt.rcParams['xtick.color'] = text_color
        plt.rcParams['ytick.color'] = text_color
        plt.rcParams['axes.edgecolor'] = grid_color
        
        # Layout: Responsive GridSpec
        from matplotlib.gridspec import GridSpec
        
        # Check current width to decide layout
        try:
            current_width = int(self.canvas_widget.winfo_width())
        except Exception:
            current_width = 0
        if current_width <= 1 and self._last_canvas_size:
            current_width = int(self._last_canvas_size[0])
        self._layout_mode = self._layout_mode_for_width(current_width)
        if self._layout_mode == "vertical":
            # Vertical Stack (Small Screen)
            gs = GridSpec(2, 1, figure=self.fig, hspace=0.4)
            ax1 = self.fig.add_subplot(gs[0, 0])
            ax2 = self.fig.add_subplot(gs[1, 0])
        else:
            # Horizontal (Large Screen)
            gs = GridSpec(1, 3, figure=self.fig, wspace=0.3)
            ax1 = self.fig.add_subplot(gs[0, :2])
            ax2 = self.fig.add_subplot(gs[0, 2])

        ax1.set_facecolor(bg_color)
        
        type_counts = df['设备缺陷类型'].value_counts().head(8) # Show more items since we have space
        if not type_counts.empty:
            # Apple Style Colors: System Blue
            bars = ax1.bar(type_counts.index, type_counts.values, color='#007AFF', width=0.6, alpha=0.9)
            ax1.set_title("缺陷类型分布 (Top 8)", fontsize=12, pad=15, color=text_color, fontweight='bold')
            ax1.tick_params(axis='x', rotation=30, labelsize=9)
            ax1.grid(axis='y', linestyle='--', alpha=0.5, color=grid_color)
            
            # Remove top and right spines for cleaner look
            ax1.spines['top'].set_visible(False)
            ax1.spines['right'].set_visible(False)
            
            # Add value labels
            for bar in bars:
                height = bar.get_height()
                ax1.text(bar.get_x() + bar.get_width()/2., height,
                        f'{int(height)}',
                        ha='center', va='bottom', color=text_color, fontsize=9)
        else:
            ax1.text(0.5, 0.5, "暂无分类数据", ha='center', va='center', color=text_color, fontsize=12)
            ax1.axis('off')
            
        total = len(df)
        closed = df['销号时间'].notna().sum()
        pending = total - closed
        
        if total > 0:
            # Apple Style Colors: Green and Red/Orange
            colors = ['#34C759', '#FF3B30'] # iOS Green, iOS Red
            wedges, texts, autotexts = ax2.pie([closed, pending], labels=['已销号', '未销号'], 
                                             autopct='%1.1f%%', colors=colors,
                                             startangle=90, pctdistance=0.85,
                                             textprops={'color': text_color, 'fontsize': 10},
                                             wedgeprops={'width': 0.4, 'edgecolor': bg_color}) # Donut style
            
            # Center text
            ax2.text(0, 0, f"{int((closed/total)*100)}%", ha='center', va='center', fontsize=14, fontweight='bold', color=text_color)
            ax2.set_title("销号完成率", fontsize=12, pad=15, color=text_color, fontweight='bold')
        else:
            ax2.text(0.5, 0.5, "暂无数据", ha='center', va='center', color=text_color, fontsize=12)
            ax2.axis('off')

        self._sync_figure_dpi_to_tk()
        try:
            self.canvas.draw_idle()
        except Exception:
            self.canvas.draw_idle()

    def export_chart(self):
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"设备缺陷统计图表_{timestamp}.png"
        
        filename = filedialog.asksaveasfilename(
            initialfile=default_filename,
            defaultextension=".png", 
            filetypes=[("PNG Image", "*.png"), ("PDF Document", "*.pdf")]
        )
        
        if filename:
            try:
                self.fig.savefig(filename)
                messagebox.showinfo("成功", f"图表已保存至: {filename}")
            except Exception as e:
                messagebox.showerror("错误", f"保存失败: {e}")

class App:
    def __init__(self, root):
        self.root = root
        # Initialize Theme with a cleaner base
        self.style = ttk.Style(theme='litera')
        
        # Configure Custom Styles for "High-End" Look
        # Card Frame: White background, subtle border
        self.style.configure("Card.TFrame", background="#FFFFFF", relief="flat", borderwidth=0)
        # Sidebar: Light gray background
        self.style.configure("Sidebar.TFrame", background="#F0F2F5")
        # Sidebar Button: Transparent by default, Pill shape active
        self.style.configure("Sidebar.TButton", background="#F0F2F5", foreground="#1D1D1F", 
                           borderwidth=0, focuscolor="#F0F2F5", font=("Microsoft YaHei UI", 14, "bold"))
        
        self.style.configure("Active.Sidebar.TButton", background="#E4E6EB", foreground="#007AFF", 
                           borderwidth=0, focuscolor="#E4E6EB", font=("Microsoft YaHei UI", 14, "bold"))

        self.style.configure("Sub.Sidebar.TButton", background="#F0F2F5", foreground="#86868B", 
                           borderwidth=0, focuscolor="#F0F2F5", font=("Microsoft YaHei UI", 12))

        self.style.configure("ActiveSub.Sidebar.TButton", background="#F0F2F5", foreground="#007AFF", 
                           borderwidth=0, focuscolor="#F0F2F5", font=("Microsoft YaHei UI", 12, "bold"))

        self.style.map("Sidebar.TButton",
                     background=[('active', '#E4E6EB'), ('selected', '#E4E6EB')],
                     foreground=[('active', '#000000'), ('selected', '#007AFF')])

        self.root.title("设备缺陷统计管理系统 V2.0")
        self.root.geometry("1200x800")
        
        # Default Maximized Window
        try:
            self.root.state('zoomed')
        except:
            self.root.attributes('-zoomed', True)

        # Logic Components
        self._app_state = _load_app_state()
        self.processor = DefectProcessor(self.log_message, self.update_progress)
        self.excel_path_var = tk.StringVar(value=self._app_state.get("excel_path") or TARGET_EXCEL_PATH)
        self._saved_source_path = self._app_state.get("source_path") or DEFAULT_SOURCE_DIR
        self._processing_lock = threading.Lock()
        self._is_processing = False
        self._cancel_reason = None
        
        # Undo/Redo/Pause state
        self.undo_stack = []
        self.redo_stack = []
        self._backup_dir = os.path.join(tempfile.gettempdir(), "defect_stats_backups")
        if not os.path.exists(self._backup_dir):
            try:
                os.makedirs(self._backup_dir, exist_ok=True)
            except Exception:
                pass

        try:
            self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        except Exception:
            pass
        
        # --- Layout: Sidebar (Left) + Content (Right) ---
        self.setup_ui()

    def setup_ui(self):
        # 1. Sidebar (Fixed width, full height)
        self.sidebar = ttk.Frame(self.root, style="Sidebar.TFrame", padding=0, width=320)
        self.sidebar.pack(side=LEFT, fill=Y)
        self.sidebar.pack_propagate(False) # Fix width
        
        # Brand Area
        brand_frame = ttk.Frame(self.sidebar, style="Sidebar.TFrame", padding=(20, 30, 20, 30))
        brand_frame.pack(fill=X)
        ttk.Label(brand_frame, text="📊 统计助手", font=("Microsoft YaHei UI", 18, "bold"), 
                 background="#F0F2F5", foreground="#333333").pack(anchor=W)
        ttk.Label(brand_frame, text="设备故障智能分析平台", font=("Microsoft YaHei UI", 9), 
                 background="#F0F2F5", foreground="#666666").pack(anchor=W, pady=(5, 0))
        
        # Navigation Menu
        self.nav_frame = ttk.Frame(self.sidebar, style="Sidebar.TFrame", padding=(10, 0))
        self.nav_frame.pack(fill=BOTH, expand=YES)
        
        self.nav_btns = {}
        self.create_nav_btn("数据采集", "collect", "📚")
        self.create_nav_btn("统计分析", "stats", "📈")
        
        # Sub-menu for Stats (Hidden by default, styled seamlessly)
        self.stats_sub_menu = ttk.Frame(self.nav_frame, style="Sidebar.TFrame")
        
        # Separator line
        sep = ttk.Frame(self.stats_sub_menu, height=2, bootstyle="secondary")
        sep.pack(fill=X, padx=0, pady=(0, 5))

        self.btn_view_chart = self.create_sub_nav_btn("图表概览", lambda: self.switch_stats_view("chart"))
        self.btn_view_list = self.create_sub_nav_btn("明细列表", lambda: self.switch_stats_view("list"))
        
        # Spacer
        ttk.Frame(self.nav_frame, style="Sidebar.TFrame").pack(fill=BOTH, expand=YES)
        
        # Bottom Actions
        bottom_frame = ttk.Frame(self.sidebar, style="Sidebar.TFrame", padding=20)
        bottom_frame.pack(side=BOTTOM, fill=X)
        
        self.create_nav_btn("关于软件", "about", "ℹ️", parent=bottom_frame)
        
        # Theme Toggle (Switch style)
        self.theme_var = tk.BooleanVar(value=False)
        self.chk_theme = ttk.Checkbutton(bottom_frame, text="深色模式", variable=self.theme_var, 
                                       command=self.toggle_theme, bootstyle="round-toggle")
        self.chk_theme.pack(anchor=W, pady=(15, 0))
        
        # 2. Main Content Container (Right)
        # Use a background color slightly different from white to show card edges
        self.content_bg = ttk.Frame(self.root) 
        self.content_bg.pack(side=RIGHT, fill=BOTH, expand=YES)
        
        # This frame holds the actual views with padding to create "Floating" effect
        self.content_container = ttk.Frame(self.content_bg, padding=20)
        self.content_container.pack(fill=BOTH, expand=YES)
        
        # Views Container
        self.views = {}
        
        # Initialize Views
        self.create_collect_view()
        self.create_stats_view()
        self.create_about_view()

        try:
            self.entry_src.delete(0, tk.END)
            self.entry_src.insert(0, self._saved_source_path)
        except Exception:
            pass
        
        # Show default
        self.show_view("collect")

    def create_nav_btn(self, text, view_name, icon="", parent=None):
        if parent is None:
            parent = self.nav_frame
            
        btn = ttk.Button(parent, text=f"  {text}", command=lambda: self.show_view(view_name),
                       style="Sidebar.TButton", cursor="hand2")
        btn.pack(pady=8, padx=12, fill=X, ipady=6)
        # Center alignment fix isn't needed with fill=X and compound, but anchor w works best
        try: btn.configure(anchor="w") 
        except: pass
            
        if view_name in ["collect", "stats", "about"]:
            self.nav_btns[view_name] = btn

    def create_sub_nav_btn(self, text, command):
        btn = ttk.Button(self.stats_sub_menu, text=f"      {text}", command=command,
                       style="Sub.Sidebar.TButton", cursor="hand2")
        btn.pack(pady=4, padx=12, fill=X, ipady=4)
        try: btn.configure(anchor="w") 
        except: pass
        return btn

    def switch_stats_view(self, view_type):
        if hasattr(self, 'stats_panel'):
            self.stats_panel.switch_view(view_type)
            # Visual feedback
            if view_type == "chart":
                self.btn_view_chart.configure(style="ActiveSub.Sidebar.TButton")
                self.btn_view_list.configure(style="Sub.Sidebar.TButton")
            else:
                self.btn_view_chart.configure(style="Sub.Sidebar.TButton")
                self.btn_view_list.configure(style="ActiveSub.Sidebar.TButton")

    def show_view(self, view_name):
        # Hide all
        for v in self.views.values():
            v.pack_forget()
        
        # Show selected
        if view_name in self.views:
            self.views[view_name].pack(fill=BOTH, expand=YES)
            
        # Toggle Sub-menu
        if view_name == "stats":
            self.stats_sub_menu.pack(after=self.nav_btns["stats"], fill=X, pady=(0, 10))
            if hasattr(self, "stats_panel"):
                def refresh_stats():
                    try:
                        self.stats_panel.load_data(silent=True)
                        self.stats_panel.request_redraw()
                    except Exception:
                        return
                self.root.after(0, refresh_stats)
        else:
            self.stats_sub_menu.pack_forget()
            
        # Update Nav State (Visual feedback)
        for name, btn in self.nav_btns.items():
            if name == view_name:
                # Active State: Highlight background or text
                btn.state(['pressed']) # Simulate pressed or use style map
                btn.configure(style="Active.Sidebar.TButton") 
            else:
                btn.configure(style="Sidebar.TButton")

    def toggle_theme(self):
        if self.theme_var.get():
            self.style.theme_use("darkly")
        else:
            self.style.theme_use("cosmo")
            
        # Update charts if they exist
        if hasattr(self, 'stats_panel'):
            self.stats_panel.render_charts()

    def on_close(self):
        try:
            state = {
                "excel_path": self.entry_dst.get() if hasattr(self, "entry_dst") else self.excel_path_var.get(),
                "source_path": self.entry_src.get() if hasattr(self, "entry_src") else DEFAULT_SOURCE_DIR,
                "saved_at": time.time(),
            }
            _save_app_state(state)
        except Exception:
            pass
        try:
            self.root.destroy()
        except Exception:
            pass

    def _create_backup(self):
        target_path = self.entry_dst.get()
        if not os.path.exists(target_path):
            return None
        
        timestamp = int(time.time() * 1000)
        backup_name = f"backup_{timestamp}.xlsx"
        backup_path = os.path.join(self._backup_dir, backup_name)
        try:
            shutil.copy2(target_path, backup_path)
            return backup_path
        except Exception as e:
            self.log_message(f"备份失败: {e}")
            return None

    def perform_undo(self):
        if not self.undo_stack:
            return
        
        if self._is_processing and not getattr(self.processor, "paused", False):
            messagebox.showwarning("提示", "正在处理中，无法撤销。")
            return

        if not messagebox.askyesno("撤销", "确定要撤销上一次操作吗？\n这将恢复Excel文件到处理前的状态。"):
            return

        try:
            if self._is_processing and getattr(self.processor, "paused", False):
                self._cancel_reason = "撤销"
                self.processor.stop_requested = True
                self.processor.paused = False
                try:
                    self.btn_pause.configure(state="disabled", text="⏸ 暂停", bootstyle="warning")
                except Exception:
                    pass

            # Save current state to redo stack
            current_backup = self._create_backup()
            if current_backup:
                self.redo_stack.append(current_backup)
            
            # Restore from undo stack
            restore_path = self.undo_stack.pop()
            target_path = self.entry_dst.get()
            
            # Ensure target directory exists just in case
            os.makedirs(os.path.dirname(target_path), exist_ok=True)
            
            shutil.copy2(restore_path, target_path)
            
            self.log_message("已撤销上一次操作。")
            self._update_action_buttons()
            
            # Refresh stats if available
            if hasattr(self, "stats_panel"):
                self.stats_panel.load_data(force=True, silent=True)
                
        except Exception as e:
            messagebox.showerror("错误", f"撤销失败: {e}")

    def perform_redo(self):
        if not self.redo_stack:
            return

        if self._is_processing and not getattr(self.processor, "paused", False):
            messagebox.showwarning("提示", "正在处理中，无法恢复。")
            return
            
        if not messagebox.askyesno("恢复", "确定要恢复撤销的操作吗？"):
            return

        try:
            if self._is_processing and getattr(self.processor, "paused", False):
                self._cancel_reason = "恢复"
                self.processor.stop_requested = True
                self.processor.paused = False
                try:
                    self.btn_pause.configure(state="disabled", text="⏸ 暂停", bootstyle="warning")
                except Exception:
                    pass

            # Save current state to undo stack
            current_backup = self._create_backup()
            if current_backup:
                self.undo_stack.append(current_backup)
                
            # Restore from redo stack
            restore_path = self.redo_stack.pop()
            target_path = self.entry_dst.get()
            shutil.copy2(restore_path, target_path)
            
            self.log_message("已恢复撤销的操作。")
            self._update_action_buttons()
            
            if hasattr(self, "stats_panel"):
                self.stats_panel.load_data(force=True, silent=True)
                
        except Exception as e:
            messagebox.showerror("错误", f"恢复失败: {e}")

    def toggle_pause(self):
        if not self._is_processing:
            return
            
        if not hasattr(self.processor, "paused"):
            self.processor.paused = False
        self.processor.paused = not self.processor.paused
        if self.processor.paused:
            self.btn_pause.configure(text="▶ 继续", bootstyle="info")
            self.log_message("已暂停处理...")
        else:
            self.btn_pause.configure(text="⏸ 暂停", bootstyle="warning")
            self.log_message("继续处理...")
        self._update_action_buttons()
            
    def _update_action_buttons(self):
        allow = (not self._is_processing) or getattr(self.processor, "paused", False)
        state_undo = "normal" if self.undo_stack and allow else "disabled"
        state_redo = "normal" if self.redo_stack and allow else "disabled"
        
        try:
            if hasattr(self, "btn_undo"):
                self.btn_undo.configure(state=state_undo)
            if hasattr(self, "btn_redo"):
                self.btn_redo.configure(state=state_redo)
        except Exception:
            pass

    def create_collect_view(self):
        view = ttk.Frame(self.content_container)
        self.views["collect"] = view
        
        # Header
        ttk.Label(view, text="数据采集与处理", font=("Microsoft YaHei UI", 18, "bold")).pack(anchor=W, pady=(10, 20))
        
        # Main Card
        card = ttk.Frame(view, style="Card.TFrame", padding=30)
        card.pack(fill=X)
        
        # Section 1: Source
        src_frame = ttk.Frame(card, style="Card.TFrame")
        src_frame.pack(fill=X, pady=(0, 20))
        
        ttk.Label(src_frame, text="数据源位置", font=("Microsoft YaHei UI", 12, "bold"), background="#FFFFFF", foreground="#000000").pack(anchor=W, pady=(0, 10))
        
        src_input_frame = ttk.Frame(src_frame, style="Card.TFrame")
        src_input_frame.pack(fill=X)
        
        self.entry_src = ttk.Entry(src_input_frame, font=("Microsoft YaHei UI", 10))
        self.entry_src.pack(side=LEFT, fill=X, expand=YES, padx=(0, 10), ipady=5)
        self.entry_src.insert(0, DEFAULT_SOURCE_DIR)
        
        ttk.Button(src_input_frame, text="📁 选择文件夹", command=self.browse_folder, bootstyle="outline-primary").pack(side=LEFT, padx=5)
        ttk.Button(src_input_frame, text="📄 选择文件", command=self.browse_file, bootstyle="outline-info").pack(side=LEFT, padx=5)

        # Section 2: Target
        dst_frame = ttk.Frame(card, style="Card.TFrame")
        dst_frame.pack(fill=X, pady=(0, 20))
        
        ttk.Label(dst_frame, text="目标 Excel 文件", font=("Microsoft YaHei UI", 12, "bold"), background="#FFFFFF", foreground="#000000").pack(anchor=W, pady=(0, 10))
        
        dst_input_frame = ttk.Frame(dst_frame, style="Card.TFrame")
        dst_input_frame.pack(fill=X)
        
        self.entry_dst = ttk.Entry(dst_input_frame, textvariable=self.excel_path_var, font=("Microsoft YaHei UI", 10))
        self.entry_dst.pack(side=LEFT, fill=X, expand=YES, padx=(0, 10), ipady=5)
        
        ttk.Button(dst_input_frame, text="📂 选择文件", command=self.browse_dst, bootstyle="outline-warning").pack(side=LEFT, padx=5)

        # Section 3: Actions & Progress
        action_frame = ttk.Frame(card, style="Card.TFrame")
        action_frame.pack(fill=X, pady=(20, 0))
        
        # Big Start Button
        self.btn_run = ttk.Button(action_frame, text="▶ 开始处理", command=self.run_process_thread, bootstyle="primary", width=20)
        self.btn_run.pack(side=LEFT, padx=(0, 15), ipady=5)
        
        self.btn_pause = ttk.Button(action_frame, text="⏸ 暂停", command=self.toggle_pause, bootstyle="warning-outline", width=10, state="disabled")
        self.btn_pause.pack(side=LEFT, padx=5, ipady=5)
        
        # Progress Bar (Modern & Thin)
        progress_frame = ttk.Frame(action_frame, style="Card.TFrame")
        progress_frame.pack(side=LEFT, fill=X, expand=YES, padx=20)
        
        self.status_var = tk.StringVar(value="准备就绪")
        ttk.Label(progress_frame, textvariable=self.status_var, background="#FFFFFF", foreground="#666666").pack(anchor=W)
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100, bootstyle="success-striped")
        self.progress_bar.pack(fill=X, pady=(5, 0))

        # Undo/Redo
        undo_frame = ttk.Frame(action_frame, style="Card.TFrame")
        undo_frame.pack(side=RIGHT)
        
        self.btn_undo = ttk.Button(undo_frame, text="↶ 撤销", command=self.perform_undo, bootstyle="secondary-outline", state="disabled", width=8)
        self.btn_undo.pack(side=LEFT, padx=5)
        self.btn_redo = ttk.Button(undo_frame, text="↷ 重做", command=self.perform_redo, bootstyle="secondary-outline", state="disabled", width=8)
        self.btn_redo.pack(side=LEFT)

        # Log Section
        log_frame = ttk.Frame(view, style="Card.TFrame", padding=20)
        log_frame.pack(fill=BOTH, expand=YES, pady=(20, 0))
        
        ttk.Label(log_frame, text="运行日志", font=("Microsoft YaHei UI", 12, "bold"), background="#FFFFFF", foreground="#000000").pack(anchor=W, pady=(0, 10))
        
        self.log_area = ScrolledText(log_frame, height=10, state='disabled', font=("Consolas", 10))
        self.log_area.pack(fill=BOTH, expand=YES)

    def create_stats_view(self):
        self.stats_panel = StatisticsPanel(self.content_container, self.excel_path_var, app_instance=self)
        self.views["stats"] = self.stats_panel

    def create_about_view(self):
        view = ttk.Frame(self.content_container)
        self.views["about"] = view
        
        # Center container
        center_frame = ttk.Frame(view)
        center_frame.pack(expand=YES, fill=BOTH, padx=20, pady=20)
        
        # Card - Make it wider and more spacious
        card = ttk.Frame(center_frame, bootstyle="secondary", padding=(40, 40))
        card.pack(anchor=CENTER, fill=X, padx=50)
        
        # Main Info Grid
        info_grid = ttk.Frame(card, bootstyle="secondary")
        info_grid.pack(fill=X, pady=(0, 0))
        info_grid.columnconfigure(0, weight=1)
        info_grid.columnconfigure(1, weight=1)

        # Unit (Row 0, Col 0)
        unit_frame = ttk.Frame(info_grid, bootstyle="secondary")
        unit_frame.grid(row=0, column=0, sticky="nw", pady=(0, 30))
        
        ttk.Label(unit_frame, text="单位", font=("Microsoft YaHei UI", 16, "bold"), bootstyle="inverse-secondary").pack(anchor=W, pady=(0, 10))
        ttk.Label(unit_frame, text="惠州电务段汕头水电车间", font=("Microsoft YaHei UI", 14), bootstyle="inverse-secondary").pack(anchor=W)

        # Technical Guidance (Row 0, Col 1)
        guide_frame = ttk.Frame(info_grid, bootstyle="secondary")
        guide_frame.grid(row=0, column=1, sticky="nw", padx=(20, 0), pady=(0, 30))
        
        ttk.Label(guide_frame, text="技术指导", font=("Microsoft YaHei UI", 16, "bold"), bootstyle="inverse-secondary").pack(anchor=W, pady=(0, 10))
        ttk.Label(guide_frame, text="李海东、梁成欧、庄金旺、郭新城、洪映森", font=("Microsoft YaHei UI", 14), bootstyle="inverse-secondary").pack(anchor=W)
        
        # Separator (Row 1)
        sep = ttk.Separator(info_grid, bootstyle="secondary")
        sep.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 30))

        # Author (Row 2, Col 0)
        author_frame = ttk.Frame(info_grid, bootstyle="secondary")
        author_frame.grid(row=2, column=0, sticky="nw")
        
        ttk.Label(author_frame, text="作者", font=("Microsoft YaHei UI", 16, "bold"), bootstyle="inverse-secondary").pack(anchor=W, pady=(0, 10))
        ttk.Label(author_frame, text="智轨先锋组", font=("Microsoft YaHei UI", 14), bootstyle="inverse-secondary").pack(anchor=W)
        
        # Contact (Row 2, Col 1)
        contact_frame = ttk.Frame(info_grid, bootstyle="secondary")
        contact_frame.grid(row=2, column=1, sticky="nw", padx=(20, 0))
        
        ttk.Label(contact_frame, text="联系方式", font=("Microsoft YaHei UI", 16, "bold"), bootstyle="inverse-secondary").pack(anchor=W, pady=(0, 10))
        
        contact_grid = ttk.Frame(contact_frame, bootstyle="secondary")
        contact_grid.pack(anchor=W)
        
        ttk.Label(contact_grid, text="电话: 19119383440", font=("Microsoft YaHei UI", 14), bootstyle="inverse-secondary").pack(side=LEFT, padx=(0, 20))
        ttk.Label(contact_grid, text="微信: yh19119383440", font=("Microsoft YaHei UI", 14), bootstyle="inverse-secondary").pack(side=LEFT)

        # --- Footer Section: Badges ---
        footer = ttk.Frame(card, bootstyle="secondary")
        footer.pack(fill=X, pady=(40, 0))
        
        badges = [("🚄 广铁定制版", "info"), ("🎨 个性化引擎", "warning"), ("🛡️ 本地存储", "primary")]
        
        # Center the badges
        badge_container = ttk.Frame(footer, bootstyle="secondary")
        badge_container.pack(anchor=CENTER)
        
        for text, style in badges:
             lbl = ttk.Label(badge_container, text=f" {text} ", bootstyle=f"{style}-inverse", font=("Microsoft YaHei UI", 12))
             lbl.pack(side=LEFT, padx=15)

    def run_sync_process_from_stats(self):
        if not messagebox.askyesno("确认", "是否开始同步流程？\n\n即将执行：\n1. 扫描并导入新增的Word文档\n2. 清理Excel中已删除文档的记录"):
            return
            
        self.run_process_thread(is_sync=True, sync_word=False)

    # --- Actions ---
    def browse_folder(self):
        initial = self.entry_src.get()
        start_dir = initial if os.path.isdir(initial) else os.path.dirname(initial) if os.path.exists(initial) else None

        d = filedialog.askdirectory(initialdir=start_dir)
        if d:
            self.entry_src.delete(0, tk.END)
            self.entry_src.insert(0, d)

    def browse_file(self):
        initial = self.entry_src.get()
        start_dir = os.path.dirname(initial) if os.path.exists(initial) else None

        f = filedialog.askopenfilename(
            initialdir=start_dir,
            filetypes=[("Word Documents", "*.doc;*.docx")]
        )
        if f:
            self.entry_src.delete(0, tk.END)
            self.entry_src.insert(0, f)

    def browse_dst(self):
        f = filedialog.askopenfilename(initialdir=os.path.dirname(self.entry_dst.get()), filetypes=[("Excel files", "*.xlsx")])
        if f:
            self.entry_dst.delete(0, tk.END)
            self.entry_dst.insert(0, f)

    def log_message(self, msg):
        self.root.after(0, self._append_log, msg)

    def _append_log(self, msg):
        try:
            # Try to use the inner text widget if available (ttkbootstrap ScrolledText)
            text_widget = getattr(self.log_area, 'text', self.log_area)
            text_widget.configure(state='normal')
            text_widget.insert(tk.END, f"[{time.strftime('%H:%M:%S')}] {msg}\n")
            text_widget.see(tk.END)
            text_widget.configure(state='disabled')
        except Exception as e:
            print(f"Log error: {e}")

    def update_progress(self, current, total, status_msg):
        self.root.after(0, self._update_ui_progress, current, total, status_msg)

    def _update_ui_progress(self, current, total, status_msg):
        if total > 0:
            pct = (current / total) * 100
            self.progress_var.set(pct)
        self.status_var.set(f"{status_msg} ({current}/{total})")

    def run_process_thread(self, is_sync=False, sync_word=False):
        src = self.entry_src.get()
        dst = self.entry_dst.get()

        with self._processing_lock:
            if self._is_processing:
                messagebox.showinfo("提示", "当前正在处理，请稍候完成后再操作。")
                return
            self._is_processing = True

        self._cancel_reason = None
        self.processor.stop_requested = False
        self.processor.paused = False
        
        # Create Backup
        backup_path = self._create_backup()
        if backup_path:
            self.undo_stack.append(backup_path)
            self.redo_stack.clear()
            self.root.after(0, self._update_action_buttons)
        
        try:
            self.btn_run.config(state='disabled')
            self.btn_pause.config(state='normal')
        except Exception:
            pass
        try:
            if hasattr(self, "stats_panel") and hasattr(self.stats_panel, "btn_sync"):
                self.stats_panel.btn_sync.config(state='disabled')
        except Exception:
            pass
        
        # Clear log
        try:
            text_widget = getattr(self.log_area, 'text', self.log_area)
            text_widget.configure(state='normal')
            text_widget.delete(1.0, tk.END)
            text_widget.configure(state='disabled')
        except Exception as e:
            print(f"Log clear error: {e}")
            
        self.progress_var.set(0)
        
        def task():
            try:
                success = self.processor.process_source(src, dst, overwrite=False, incremental=is_sync)
                
                if success and sync_word:
                    self.processor.log("开始执行反向同步...")
                    success = self.processor.sync_word_from_excel(dst)

                if success:
                    if is_sync:
                        self.root.after(0, lambda: self.stats_panel.load_data(force=True, silent=True))
                        self.root.after(0, lambda: messagebox.showinfo("完成", "同步完成！"))
                    else:
                        self.root.after(0, lambda: messagebox.showinfo("完成", "数据采集处理完成！\n请切换到“统计分析”查看结果。"))
                else:
                    if self._cancel_reason:
                        msg = f"已{self._cancel_reason}，并停止当前处理。"
                        self.root.after(0, lambda m=msg: messagebox.showinfo("已停止", m))
                    else:
                        self.root.after(0, lambda: messagebox.showerror("失败", "处理过程中遇到错误，请检查日志。"))
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("异常", str(e)))
            finally:
                def finish():
                    with self._processing_lock:
                        self._is_processing = False
                    try:
                        self.btn_run.config(state='normal')
                        self.btn_pause.configure(state="disabled", text="⏸ 暂停", bootstyle="warning")
                        self.processor.paused = False
                        self.processor.stop_requested = False
                    except Exception:
                        pass
                    try:
                        if hasattr(self, "stats_panel") and hasattr(self.stats_panel, "btn_sync"):
                            self.stats_panel.btn_sync.config(state='normal')
                    except Exception:
                        pass
                    try:
                        self.status_var.set("就绪")
                    except Exception:
                        pass
                    self._cancel_reason = None
                    self._update_action_buttons()

                self.root.after(0, finish)

        threading.Thread(target=task, daemon=True).start()

if __name__ == "__main__":
    import ttkbootstrap as ttk
    root = ttk.Window(themename="cosmo")
    app = App(root)
    root.mainloop()
