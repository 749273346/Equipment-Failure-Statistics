import os
import tempfile
import tkinter as tk

import openpyxl
import ttkbootstrap as ttk

import auto_fill_defects as afd


def test_excel_write_rows():
    with tempfile.TemporaryDirectory() as d:
        excel_path = os.path.join(d, "test.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=3, column=1, value="1")
        ws.cell(row=3, column=2, value="模板")
        wb.save(excel_path)

        long_path = os.path.join(d, "a" * 80, "b" * 80, "demo.docx")
        row_data = ["", "广州", "类型A", "地点A", "", "", "", "", "", "", "", "", "", long_path]
        row_data2 = ["", "深圳", "类型B", "地点B", "", "", "", "", "", "", "", "", "", long_path]

        p = afd.DefectProcessor(log_callback=lambda *_: None)
        wrote = p._write_rows_to_excel(excel_path, [row_data])
        if wrote != 1:
            raise RuntimeError(f"写入条数异常: {wrote}")

        wrote2 = p._write_rows_to_excel(excel_path, [row_data2])
        if wrote2 != 1:
            raise RuntimeError(f"追加写入条数异常: {wrote2}")

        wb2 = openpyxl.load_workbook(excel_path)
        ws2 = wb2.active

        if ws2.max_row != 5:
            raise RuntimeError(f"行数异常: {ws2.max_row}")

        cell_path = ws2.cell(row=4, column=14).value
        if str(cell_path) != long_path:
            raise RuntimeError("源文件路径未写入第14列")

        if not ws2.column_dimensions["N"].hidden:
            raise RuntimeError("第14列未隐藏")

        h = ws2.row_dimensions[4].height
        if h is not None and h > 150:
            raise RuntimeError(f"行高异常: {h}")

        wrote_refresh = p._write_rows_to_excel(excel_path, [row_data, row_data2], overwrite=True)
        if wrote_refresh != 2:
            raise RuntimeError(f"刷新写入条数异常: {wrote_refresh}")

        wb3 = openpyxl.load_workbook(excel_path)
        ws3 = wb3.active
        if ws3.max_row != 5:
            raise RuntimeError(f"刷新后行数异常: {ws3.max_row}")
        if str(ws3.cell(row=4, column=1).value) != "1":
            raise RuntimeError("刷新后序号异常")


def test_undo_redo_pause():
    with tempfile.TemporaryDirectory() as d:
        excel_path = os.path.join(d, "undo_redo.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "before"
        wb.save(excel_path)

        afd.messagebox.askyesno = lambda *args, **kwargs: True
        afd.messagebox.showwarning = lambda *args, **kwargs: None
        afd.messagebox.showerror = lambda *args, **kwargs: None
        afd.messagebox.showinfo = lambda *args, **kwargs: None

        root = ttk.Window(themename="cosmo")
        try:
            root.withdraw()
        except Exception:
            pass

        app = afd.App(root)
        app.stats_panel.load_data = lambda *args, **kwargs: None
        if not hasattr(app.processor, "paused"):
            raise RuntimeError("DefectProcessor 初始化未包含 paused 字段")
        if app.processor.paused is not False:
            raise RuntimeError("DefectProcessor.paused 初始值应为 False")

        app.entry_dst.delete(0, tk.END)
        app.entry_dst.insert(0, excel_path)

        backup = app._create_backup()
        if not backup or not os.path.exists(backup):
            raise RuntimeError("初始备份创建失败")
        app.undo_stack.append(backup)
        app.redo_stack.clear()
        app._update_action_buttons()

        wb2 = openpyxl.load_workbook(excel_path)
        ws2 = wb2.active
        ws2["A1"] = "after"
        wb2.save(excel_path)

        app.perform_undo()
        wb3 = openpyxl.load_workbook(excel_path)
        if wb3.active["A1"].value != "before":
            raise RuntimeError("撤销未恢复到上次导入前状态")
        if app.redo_stack == []:
            raise RuntimeError("撤销后redo栈为空")

        app.perform_redo()
        wb4 = openpyxl.load_workbook(excel_path)
        if wb4.active["A1"].value != "after":
            raise RuntimeError("恢复未恢复到撤销前状态")
        if app.redo_stack:
            raise RuntimeError("恢复后redo栈未清空")

        app._is_processing = True
        app.btn_pause.configure(state="normal")
        app.processor.paused = False
        app.toggle_pause()
        if not app.processor.paused:
            raise RuntimeError("暂停未生效")
        if app.btn_pause.cget("text") != "▶ 继续":
            raise RuntimeError("暂停按钮文字未切换到继续")

        if not app.undo_stack:
            backup_paused = app._create_backup()
            if not backup_paused or not os.path.exists(backup_paused):
                raise RuntimeError("暂停场景初始撤销栈为空且备份创建失败")
            app.undo_stack.append(backup_paused)

        app._update_action_buttons()
        if app.btn_undo.instate(["disabled"]):
            raise RuntimeError("暂停状态下撤销按钮应可用")

        backup2 = app._create_backup()
        if not backup2 or not os.path.exists(backup2):
            raise RuntimeError("暂停场景备份创建失败")
        app.undo_stack.append(backup2)

        wb5 = openpyxl.load_workbook(excel_path)
        ws5 = wb5.active
        ws5["A1"] = "after2"
        wb5.save(excel_path)

        app.perform_undo()
        wb6 = openpyxl.load_workbook(excel_path)
        if wb6.active["A1"].value != "after":
            raise RuntimeError("暂停状态下撤销未恢复到上次状态")
        if app.processor.paused:
            raise RuntimeError("暂停状态下执行撤销后应自动解除暂停")
        if not app.btn_pause.instate(["disabled"]):
            raise RuntimeError("暂停状态下执行撤销后应禁用暂停按钮以停止当前处理")

        try:
            root.destroy()
        except Exception:
            pass


def test_filtering_year_month_status():
    afd.messagebox.askyesno = lambda *args, **kwargs: True
    afd.messagebox.showwarning = lambda *args, **kwargs: None
    afd.messagebox.showerror = lambda *args, **kwargs: None
    afd.messagebox.showinfo = lambda *args, **kwargs: None

    root = ttk.Window(themename="cosmo")
    try:
        root.withdraw()
    except Exception:
        pass

    excel_var = tk.StringVar(value="")
    panel = afd.StatisticsPanel(root, excel_var, app_instance=None)

    df = afd.pd.DataFrame(
        {
            "序号": [1, 2, 3, 4],
            "设备缺陷地点": ["A", "B", "C", "D"],
            "设备缺陷类型": ["T1", "T2", "T3", "T4"],
            "发现时间": ["2024-06-01", "2025-01-10", "2025-02-01", "2024-03-01"],
            "销号时间": ["2024-12-13", "2025-01-20", None, None],
        }
    )

    panel.df = df
    panel.year_var.set("2025")
    panel.month_var.set("全部")
    panel.status_filter_var.set("全部状态")
    panel.search_var.set("")

    out = panel.filter_dataframe(panel.df, True, True, False)
    got = set(out["序号"].tolist())
    if got != {2, 3}:
        raise RuntimeError(f"年份=2025 且 状态=全部状态 结果异常: {got}")

    panel.status_filter_var.set("未销号")
    out = panel.filter_dataframe(panel.df, True, True, False)
    got = set(out["序号"].tolist())
    if got != {3}:
        raise RuntimeError(f"年份=2025 且 状态=未销号 结果异常: {got}")

    panel.year_var.set("2024")
    panel.status_filter_var.set("未销号")
    out = panel.filter_dataframe(panel.df, True, True, False)
    got = set(out["序号"].tolist())
    if got != {4}:
        raise RuntimeError(f"年份=2024 且 状态=未销号 结果异常: {got}")

    panel.status_filter_var.set("已销号")
    out = panel.filter_dataframe(panel.df, True, True, False)
    got = set(out["序号"].tolist())
    if got != {1}:
        raise RuntimeError(f"年份=2024 且 状态=已销号 结果异常: {got}")

    panel.year_var.set("2025")
    panel.month_var.set("2月")
    panel.status_filter_var.set("全部状态")
    out = panel.filter_dataframe(panel.df, True, True, False)
    got = set(out["序号"].tolist())
    if got != {3}:
        raise RuntimeError(f"年份=2025 且 月份=2月 且 状态=全部状态 结果异常: {got}")

    try:
        root.destroy()
    except Exception:
        pass


def main():
    test_excel_write_rows()
    test_undo_redo_pause()
    test_filtering_year_month_status()
    print("OK")


if __name__ == "__main__":
    main()
